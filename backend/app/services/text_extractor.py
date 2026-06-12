"""Text extraction service for supported file formats.

Extracts plain text content from files based on their format.
Supports: .txt, .md, .json, .docx, .pdf, .csv, .xlsx, .xls
Returns None for unsupported formats or extraction failures.
"""

import csv
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class TextExtractor:
    """Extracts text content from files based on their format."""

    # Mapping of file extensions to extraction methods
    _SUPPORTED_FORMATS = {".txt", ".md", ".json", ".docx", ".pdf", ".csv", ".xlsx", ".xls"}

    def extract(self, file_path: Path, file_format: str) -> str | None:
        """Extract text from a file.

        Args:
            file_path: Path to the file on disk.
            file_format: The file extension (e.g., ".txt", ".md").

        Returns:
            Extracted text content, or None if the format is unsupported
            or extraction fails.
        """
        # Normalize the format to lowercase with leading dot
        fmt = file_format.lower() if file_format.startswith(".") else f".{file_format.lower()}"

        if fmt not in self._SUPPORTED_FORMATS:
            logger.warning(
                f"Unsupported file format '{fmt}' for file: {file_path}"
            )
            return None

        try:
            extractor = {
                ".txt": self._extract_plaintext,
                ".md": self._extract_markdown,
                ".json": self._extract_json,
                ".docx": self._extract_docx,
                ".pdf": self._extract_pdf,
                ".csv": self._extract_csv,
                ".xlsx": self._extract_xlsx,
                ".xls": self._extract_xlsx,
            }[fmt]
            return extractor(file_path)
        except Exception as e:
            logger.error(
                f"Failed to extract text from '{file_path}' (format: {fmt}): {e}"
            )
            return None

    def _extract_plaintext(self, path: Path) -> str:
        """Extract text from a plain text file.

        Args:
            path: Path to the .txt file.

        Returns:
            The raw text content of the file.
        """
        return path.read_text(encoding="utf-8")

    def _extract_markdown(self, path: Path) -> str:
        """Extract text from a Markdown file.

        Markdown is treated as plain text since the markup is
        semantically meaningful for embedding purposes.

        Args:
            path: Path to the .md file.

        Returns:
            The raw Markdown content of the file.
        """
        return path.read_text(encoding="utf-8")

    def _extract_json(self, path: Path) -> str:
        """Extract text from a JSON file by pretty-printing its content.

        Args:
            path: Path to the .json file.

        Returns:
            Pretty-printed JSON string representation.
        """
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return json.dumps(data, indent=2, ensure_ascii=False)

    def _extract_docx(self, path: Path) -> str:
        """Extract text from a DOCX file using python-docx.

        Extracts text from all paragraphs in the document,
        joining them with newlines.

        Args:
            path: Path to the .docx file.

        Returns:
            Concatenated paragraph text from the document.
        """
        from docx import Document

        doc = Document(str(path))
        paragraphs = [paragraph.text for paragraph in doc.paragraphs]
        return "\n".join(paragraphs)

    def _extract_pdf(self, path: Path) -> str:
        try:
            import fitz
            doc = fitz.open(str(path))
            pages = [page.get_text() for page in doc]
            doc.close()
            return "\n".join(pages)
        except Exception:
            return path.read_text(encoding="utf-8")

    def _extract_xlsx(self, path: Path) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            rows = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_rows = []
                for row in ws.iter_rows(values_only=True):
                    # Filter out None/empty cells to reduce noise
                    cleaned = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if cleaned:
                        sheet_rows.append(" | ".join(cleaned))
                if sheet_rows:
                    rows.append(f"--- Sheet: {sheet_name} ---")
                    rows.extend(sheet_rows)
            wb.close()
            return "\n".join(rows)
        except Exception:
            return path.read_text(encoding="utf-8")

    def _extract_csv(self, path: Path) -> str:
        """Extract text from a CSV file as formatted rows.

        Args:
            path: Path to the .csv file.

        Returns:
            Tab-separated text representation of the CSV data.
        """
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = [" | ".join(row) for row in reader]
        return "\n".join(rows)
