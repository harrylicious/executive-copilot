"""Structured data extraction from multi-sheet Excel workbooks."""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


@dataclass
class ParsedSheet:
    """A single parsed worksheet."""

    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]
    row_count: int
    data_types: dict[str, str]  # {column_name: "string"|"numeric"|"date"}


@dataclass
class ParsedWorkbook:
    """Complete parsed workbook output."""

    file_path: str
    sheets: list[ParsedSheet] = field(default_factory=list)
    parse_errors: list[str] = field(default_factory=list)


class FileSizeExceededError(Exception):
    """Raised when an Excel file exceeds the maximum allowed size."""

    def __init__(self, file_path: str, size_mb: float, max_size_mb: int) -> None:
        self.file_path = file_path
        self.size_mb = size_mb
        self.max_size_mb = max_size_mb
        super().__init__(
            f"File '{file_path}' is {size_mb:.2f} MB, exceeding the "
            f"{max_size_mb} MB limit."
        )


class ExcelParseError(Exception):
    """Raised when an Excel file cannot be opened or has no readable sheets."""

    def __init__(self, file_path: str, reason: str) -> None:
        self.file_path = file_path
        self.reason = reason
        super().__init__(f"Cannot parse '{file_path}': {reason}")


class StructuredDataExtractor:
    """Parses Excel workbooks into typed relational structures."""

    MAX_FILE_SIZE_MB: int = 50
    MAX_SHEETS: int = 50
    MAX_ROWS_PER_SHEET: int = 100_000
    MAX_COLS_PER_SHEET: int = 500

    def extract(self, file_path: Path) -> ParsedWorkbook:
        """Parse an Excel file into structured sheet data.

        Raises:
            FileSizeExceededError: If file exceeds 50 MB.
            ExcelParseError: If file cannot be opened or has no readable sheets.
        """
        file_path = Path(file_path)

        # Validate file exists and get size
        try:
            file_size_bytes = file_path.stat().st_size
        except (OSError, FileNotFoundError) as e:
            raise ExcelParseError(str(file_path), str(e)) from e

        # Validate file size
        file_size_mb = file_size_bytes / (1024 * 1024)
        if file_size_mb > self.MAX_FILE_SIZE_MB:
            raise FileSizeExceededError(
                str(file_path), file_size_mb, self.MAX_FILE_SIZE_MB
            )

        # Open workbook
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True
            )
        except Exception as e:
            raise ExcelParseError(str(file_path), str(e)) from e

        # Parse sheets
        parsed_workbook = ParsedWorkbook(file_path=str(file_path))
        sheet_names = workbook.sheetnames[: self.MAX_SHEETS]

        for sheet_name in sheet_names:
            try:
                worksheet = workbook[sheet_name]
                parsed_sheet = self._parse_sheet(worksheet, sheet_name)
                if parsed_sheet is not None:
                    parsed_workbook.sheets.append(parsed_sheet)
            except Exception as e:
                parsed_workbook.parse_errors.append(
                    f"Error parsing sheet '{sheet_name}': {e}"
                )

        workbook.close()

        if not parsed_workbook.sheets and not parsed_workbook.parse_errors:
            raise ExcelParseError(str(file_path), "No readable sheets found")

        return parsed_workbook

    def _parse_sheet(self, worksheet, sheet_name: str) -> ParsedSheet | None:
        """Parse a single worksheet into typed records."""
        # Read all rows from the worksheet (up to limits)
        all_rows: list[list[Any]] = []
        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            if i >= self.MAX_ROWS_PER_SHEET:
                break
            # Truncate columns to limit
            row_data = list(row[: self.MAX_COLS_PER_SHEET])
            all_rows.append(row_data)

        if not all_rows:
            return None

        # Filter empty rows and columns to preserve contiguous data region
        all_rows, col_indices = self._filter_empty_rows_and_cols(all_rows)

        if not all_rows or not col_indices:
            return None

        # Detect header row — scan up to the first 10 rows to skip title/
        # metadata rows that appear before the real column header row.
        # Strategy: pick the candidate row with the most non-empty text
        # (non-numeric) cells. Title rows typically span 1–2 cells;
        # type-descriptor rows ("Text", "Whole Number") are all text but
        # their values are generic data-type words; the real header row
        # has the highest count of *distinct* text values that are not
        # common Excel type-descriptor words.
        _TYPE_DESCRIPTOR_WORDS = {
            "text", "number", "whole number", "decimal number", "date",
            "true/false", "currency", "percentage", "scientific",
        }
        header_row_idx = 0
        max_scan = min(10, len(all_rows))
        best_score = -1
        for candidate_idx in range(max_scan):
            candidate = all_rows[candidate_idx]
            non_empty = [v for v in candidate if not self._is_cell_empty(v)]
            if not non_empty:
                continue
            # Count text (non-numeric) cells that are not generic type words
            meaningful_text_count = sum(
                1 for v in non_empty
                if (
                    not (isinstance(v, (int, float)) and not isinstance(v, bool))
                    and str(v).strip().lower() not in _TYPE_DESCRIPTOR_WORDS
                )
            )
            # Also reward rows with more total non-empty cells (wider rows)
            score = meaningful_text_count * 100 + len(non_empty)
            if score > best_score:
                best_score = score
                header_row_idx = candidate_idx

        is_header, headers = self._detect_header_row(all_rows[header_row_idx])

        # Deduplicate headers
        headers = self._deduplicate_headers(headers)

        # Determine data rows (everything after the detected header row)
        if is_header:
            data_rows = all_rows[header_row_idx + 1:]
        else:
            data_rows = all_rows[header_row_idx:]

        # Build typed records
        rows: list[dict[str, Any]] = []
        for row_data in data_rows:
            record: dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row_data):
                    record[header] = row_data[col_idx]
                else:
                    record[header] = None
            rows.append(record)

        # Infer data types per column
        data_types: dict[str, str] = {}
        for header in headers:
            col_values = [row.get(header) for row in rows]
            data_types[header] = self._infer_data_type(col_values)

        return ParsedSheet(
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            row_count=len(rows),
            data_types=data_types,
        )

    def _filter_empty_rows_and_cols(
        self, rows: list[list[Any]]
    ) -> tuple[list[list[Any]], list[int]]:
        """Filter empty rows and columns, preserving contiguous data region.

        Removes rows where all cells are None or whitespace-only strings,
        and removes columns where all cells are None or whitespace-only strings.
        """
        if not rows:
            return [], []

        # First pass: identify non-empty rows
        non_empty_rows: list[list[Any]] = []
        for row in rows:
            if self._is_row_non_empty(row):
                non_empty_rows.append(row)

        if not non_empty_rows:
            return [], []

        # Determine max column count
        max_cols = max(len(row) for row in non_empty_rows)

        # Second pass: identify non-empty columns
        non_empty_col_indices: list[int] = []
        for col_idx in range(max_cols):
            col_has_data = False
            for row in non_empty_rows:
                if col_idx < len(row):
                    cell = row[col_idx]
                    if not self._is_cell_empty(cell):
                        col_has_data = True
                        break
            if col_has_data:
                non_empty_col_indices.append(col_idx)

        if not non_empty_col_indices:
            return [], []

        # Build filtered result keeping only non-empty columns
        filtered_rows: list[list[Any]] = []
        for row in non_empty_rows:
            filtered_row = []
            for col_idx in non_empty_col_indices:
                if col_idx < len(row):
                    filtered_row.append(row[col_idx])
                else:
                    filtered_row.append(None)
            filtered_rows.append(filtered_row)

        return filtered_rows, non_empty_col_indices

    def _is_row_non_empty(self, row: list[Any]) -> bool:
        """Check if a row has at least one non-empty cell."""
        for cell in row:
            if not self._is_cell_empty(cell):
                return True
        return False

    def _is_cell_empty(self, cell: Any) -> bool:
        """Check if a cell value is empty (None or whitespace-only string)."""
        if cell is None:
            return True
        if isinstance(cell, str) and cell.strip() == "":
            return True
        return False

    def _detect_header_row(
        self, first_row_values: list[Any]
    ) -> tuple[bool, list[str]]:
        """Detect header row; auto-generate Column_N if no text header found.

        Returns:
            A tuple of (is_header_row, headers).
            If the first row contains only numeric values, returns False
            with auto-generated Column_1, Column_2, ... headers.
            Otherwise, returns True with the first row values as headers.
        """
        # Check if all non-empty values in the first row are numeric
        non_empty_values = [
            v for v in first_row_values if not self._is_cell_empty(v)
        ]

        if not non_empty_values:
            # All empty - generate headers
            col_count = len(first_row_values)
            headers = [f"Column_{i + 1}" for i in range(col_count)]
            return False, headers

        all_numeric = all(
            isinstance(v, (int, float)) and not isinstance(v, bool)
            for v in non_empty_values
        )

        if all_numeric:
            # First row is data, not headers - auto-generate
            col_count = len(first_row_values)
            headers = [f"Column_{i + 1}" for i in range(col_count)]
            return False, headers

        # Use first row as headers
        headers: list[str] = []
        for i, val in enumerate(first_row_values):
            if self._is_cell_empty(val):
                headers.append(f"Column_{i + 1}")
            else:
                headers.append(str(val).strip())

        return True, headers

    def _infer_data_type(self, values: list[Any]) -> str:
        """Infer column data type from cell values.

        Checks the majority type among non-empty values:
        - "numeric" if majority are int/float
        - "date" if majority are datetime
        - "string" as default
        """
        if not values:
            return "string"

        # Filter out empty values
        non_empty = [v for v in values if not self._is_cell_empty(v)]

        if not non_empty:
            return "string"

        numeric_count = 0
        date_count = 0
        string_count = 0

        for val in non_empty:
            if isinstance(val, bool):
                string_count += 1
            elif isinstance(val, (int, float)):
                numeric_count += 1
            elif isinstance(val, datetime):
                date_count += 1
            else:
                string_count += 1

        # Determine majority type
        total = len(non_empty)
        if numeric_count > total / 2:
            return "numeric"
        if date_count > total / 2:
            return "date"
        return "string"

    def _deduplicate_headers(self, headers: list[str]) -> list[str]:
        """Append numeric suffix to duplicate column names.

        For duplicate names, appends _1, _2, etc. to subsequent occurrences.
        """
        seen: dict[str, int] = {}
        result: list[str] = []

        for header in headers:
            lower_header = header.lower()
            if lower_header in seen:
                seen[lower_header] += 1
                result.append(f"{header}_{seen[lower_header]}")
            else:
                seen[lower_header] = 0
                result.append(header)

        return result
