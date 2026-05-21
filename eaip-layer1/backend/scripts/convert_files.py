"""Convert plain-text stub files to real PDF (.pdf) and Excel (.xlsx) files."""

import re
import csv
import json
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent / "knowledge_base"


def parse_ascii_table(lines):
    """Parse ASCII table (lines with | or tab separators) into list of rows."""
    rows = []
    for line in lines:
        line = line.strip()
        if not line or re.match(r'^[\s\|\-=:+]+$', line):
            continue
        if '|' in line:
            rows.append([c.strip() for c in line.split('|')])
        elif '\t' in line:
            rows.append([c.strip() for c in line.split('\t')])
    return rows


def convert_pdf(path: Path):
    import fitz
    text = path.read_text(encoding="utf-8")
    lines = text.split("\n")
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    x0, y0, w, h = 50, 50, 495, 750
    y = y0
    for line in lines:
        if y > y0 + h - 20:
            page = doc.new_page(width=595, height=842)
            y = y0
        page.insert_text((x0 + 10, y), line, fontsize=8, fontname="helv")
        y += 12
    doc.save(str(path))
    doc.close()
    print(f"  PDF: {path.relative_to(BASE)} ({len(text)} chars)")


def convert_xlsx(path: Path):
    import openpyxl
    text = path.read_text(encoding="utf-8")
    lines = text.split("\n")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    current_sheet = ws
    row_idx = 1
    for line in lines:
        if line.startswith("--- Sheet:"):
            sheet_name = line.replace("--- Sheet:", "").replace("---", "").strip()
            if sheet_name:
                ws = wb.create_sheet(title=sheet_name)
                row_idx = 1
                continue
        if not line.strip():
            row_idx += 1
            continue
        cells = [c.strip() for c in line.split("\t")] if "\t" in line else [c.strip() for c in line.split("|") if c.strip() or "|" in line]
        if not cells or all(not c for c in cells):
            row_idx += 1
            continue
        for col_idx, cell in enumerate(cells, 1):
            ws.cell(row=row_idx, column=col_idx, value=cell)
        row_idx += 1
    wb.save(str(path))
    wb.close()
    print(f"  XLSX: {path.relative_to(BASE)} ({len(text)} chars)")


def main():
    pdfs = sorted(BASE.rglob("*.pdf"))
    xlsxs = sorted(BASE.rglob("*.xlsx"))
    print("Converting PDFs...")
    for p in pdfs:
        convert_pdf(p)
    print(f"  {len(pdfs)} PDFs done")
    print("Converting XLSXs...")
    for p in xlsxs:
        convert_xlsx(p)
    print(f"  {len(xlsxs)} XLSXs done")
    print("All conversions complete!")


if __name__ == "__main__":
    main()
